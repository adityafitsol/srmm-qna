#!/usr/bin/env python3
"""
SRMM Questionnaire Scraper  —  Gemini 2.5 Pro (Vertex AI)
==========================================================
Reads companies.json, downloads each BRSR PDF, asks Gemini to score every
SRMM v2.0 indicator, calculates maturity level, and saves to:
  • srmm-qna.json   (all answers + scores, company-wise)
  • srmm-qna.xlsx   (summary + per-indicator breakdown)

Run overnight:  python3 srmm_scraper.py
Resume safely:  re-run the same command – already-done companies are skipped.
"""

import json, os, sys, time, logging, hashlib, traceback, re, signal
import requests
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import vertexai
from vertexai.generative_models import GenerativeModel, Part, GenerationConfig
from google.oauth2 import service_account
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
BASE_DIR          = Path(__file__).parent
KEY_FILE          = BASE_DIR / "quiet-mechanic-451307-s9-1bd5db312124.json"
COMPANIES_JSON    = BASE_DIR / "companies.json"
OUTPUT_JSON       = BASE_DIR / "srmm-qna.json"
OUTPUT_EXCEL      = BASE_DIR / "srmm-qna.xlsx"
PROGRESS_FILE     = BASE_DIR / ".srmm_progress.json"   # tracks done sno's

PROJECT_ID        = "quiet-mechanic-451307-s9"
LOCATION          = "us-central1"
MODEL_NAME        = "gemini-2.5-pro"

MAX_WORKERS       = 3        # parallel Gemini calls (keep low to respect quota)
RETRY_ATTEMPTS    = 3        # retries per company on transient errors
RETRY_DELAY_BASE  = 15       # seconds base for exponential backoff
REQUEST_TIMEOUT   = 300      # PDF download timeout (seconds)
PDF_DOWNLOAD_TIMEOUT = 60

# ─── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(BASE_DIR / "srmm_scraper.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# SRMM v2.0 FULL QUESTION SCHEMA
# Each entry: id, section, principle (if C), category (essential/leadership),
#             parameter, max_score, scoring_criteria (for prompt)
# ═══════════════════════════════════════════════════════════════════════════════
SRMM_SCHEMA = [
    # ── SECTION A ────────────────────────────────────────────────────────────
    {"id":"A18","sec":"A","cat":"essential","max":6,
     "param":"Categories of Employees and Workmen (differently abled %)",
     "criteria":"Score 3 if >5% of employees are differently abled; 2 for 5-2%; 1 for <2%; 0 if not engaging. Same sub-scores for workmen. Max 6 total."},
    {"id":"A19","sec":"A","cat":"essential","max":2,
     "param":"Women Employees (% of total)",
     "criteria":"Score 2 if >25%; 1 if 25-10%; 0 if <10%."},
    {"id":"A20","sec":"A","cat":"essential","max":2,
     "param":"Turnover Rate for permanent employees and workers (3-year avg)",
     "criteria":"Score 2 if avg turnover <10%; 1 if 10-15%; 0 if >15%."},
    {"id":"A21a","sec":"A","cat":"essential","max":1,
     "param":"Subsidiary Companies participating in BR Initiatives",
     "criteria":"Score 1 if Reported; 0 if Not Reported."},
    {"id":"A23","sec":"A","cat":"essential","max":4,
     "param":"Stakeholder Complaints / Grievances on Responsible Business Conduct",
     "criteria":"Score 1 for grievance/redressal mechanism in place; then 3 if no complaints or >80% resolved; 2 for 60-80%; 1 for <60%; 0 if NR. Max 4."},
    {"id":"A24","sec":"A","cat":"essential","max":3,
     "param":"Risk Assessment of ESG matters",
     "criteria":"Score 3 if assessment done + efforts to address + financial implications; 2 if only assessment; 1 for partial; 0 if NR."},
    # ── SECTION B ────────────────────────────────────────────────────────────
    {"id":"B1a","sec":"B","cat":"essential","max":3,
     "param":"Company policy covering NGRBC principles",
     "criteria":"Score 3 for all 9 principles; 2 for >5; 1 for >3; 0 if NR."},
    {"id":"B2","sec":"B","cat":"essential","max":1,
     "param":"Policy translated into procedures",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"B3","sec":"B","cat":"essential","max":1,
     "param":"Policies extended to value chain partners",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"B4","sec":"B","cat":"essential","max":3,
     "param":"National/international certifications/standards adopted and mapped to principles",
     "criteria":"Score 3 for all principles; 2 for >5; 1 for >3; 0 if NR."},
    {"id":"B5","sec":"B","cat":"essential","max":1,
     "param":"Commitments, goals and targets with defined timelines",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"B6","sec":"B","cat":"essential","max":3,
     "param":"Performance against commitments/goals/targets",
     "criteria":"Score 3 if >80% goals met; 2 for 60-80%; 1 for 50-60%; 0 for <50%."},
    {"id":"B7","sec":"B","cat":"essential","max":2,
     "param":"Director's statement on ESG challenges, targets and achievements",
     "criteria":"Score 2 for full statement with challenges/targets/achievements; 1 for partial; 0 if NR."},
    {"id":"B8","sec":"B","cat":"essential","max":1,
     "param":"Highest authority responsible for BR Policy",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"B9","sec":"B","cat":"essential","max":1,
     "param":"Board committee for sustainability decision-making",
     "criteria":"Score 1 if Yes and composition specified; 0 if No."},
    {"id":"B10","sec":"B","cat":"essential","max":3,
     "param":"Review frequency of NGRBCs",
     "criteria":"Score 3 for half-yearly review; 2 for annual; 0 for no review."},
    {"id":"B11","sec":"B","cat":"essential","max":5,
     "param":"External independent assessment/audit of policies",
     "criteria":"Score 5 for external assessment/audit; 0 if no assessment."},
    # ── PRINCIPLE 1 ──────────────────────────────────────────────────────────
    {"id":"P1E_1.1a","sec":"C","pri":"P1","cat":"essential","max":3,
     "param":"Training and awareness programs on NGRBC principles",
     "criteria":"Score 3 for Directors+KMP+employees+others; 2 for any two; 1 for one; 0 if none/NR."},
    {"id":"P1E_1.1b","sec":"C","pri":"P1","cat":"essential","max":5,
     "param":"% coverage by awareness programs on principles",
     "criteria":"Score 5 for >90%; 4 for 80-90%; 3 for 60-80%; 2 for 50-60%; 0 for <50% or no program."},
    {"id":"P1E_1.2a","sec":"C","pri":"P1","cat":"essential","max":1,
     "param":"Fines/penalties/settlement paid in financial year",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P1E_1.2b","sec":"C","pri":"P1","cat":"essential","max":1,
     "param":"Monetary Penalty/Fine/Award/Compounding fee details",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P1E_1.2c","sec":"C","pri":"P1","cat":"essential","max":1,
     "param":"Non-Monetary Cases (Imprisonment/Punishment)",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P1E_1.3","sec":"C","pri":"P1","cat":"essential","max":1,
     "param":"Appeal/Revision in cases where penalties impugned",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P1E_1.4","sec":"C","pri":"P1","cat":"essential","max":3,
     "param":"Anti-corruption / anti-bribery policy",
     "criteria":"Score 3 for policy exists and reported; 1 for no policy but reported; 0 if NR."},
    {"id":"P1E_1.5","sec":"C","pri":"P1","cat":"essential","max":1,
     "param":"Directors/KMP/Employees disciplinary action for bribery/corruption",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P1E_1.6","sec":"C","pri":"P1","cat":"essential","max":1,
     "param":"Complaints on Conflict of Interest of Directors/KMP",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P1E_1.7","sec":"C","pri":"P1","cat":"essential","max":2,
     "param":"Corrective action on corruption/conflict of interest issues",
     "criteria":"Score 2 if reported and corrective actions taken; 1 if reported only; 0 if NR."},
    {"id":"P1L_L1.1","sec":"C","pri":"P1","cat":"leadership","max":3,
     "param":"Awareness programs for value chain partners (P1)",
     "criteria":"Score 3 for all partners all principles; 2 for >2 partners >5 principles; 1 for 1-2 partners; 0 otherwise."},
    {"id":"P1L_L1.2","sec":"C","pri":"P1","cat":"leadership","max":2,
     "param":"Processes to manage Board conflict of interest",
     "criteria":"Score 2 if process exists and reported; 1 if no process but reported; 0 if NR."},
    # ── PRINCIPLE 2 ──────────────────────────────────────────────────────────
    {"id":"P2E_2.1","sec":"C","pri":"P2","cat":"essential","max":5,
     "param":"% of R&D and capex for environmental/social impact improvement",
     "criteria":"Score 5 for >40%; 4 for >30%; 3 for >20%; 2 for 10-20%; 1 for <10%; 0 if NR."},
    {"id":"P2E_2.2a","sec":"C","pri":"P2","cat":"essential","max":1,
     "param":"Sustainable sourcing procedures in place",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"P2E_2.2b","sec":"C","pri":"P2","cat":"essential","max":5,
     "param":"% of inputs sourced sustainably",
     "criteria":"Score 5 for >75%; 4 for 50-75%; 3 for 25-50%; 2 for 10-25%; 1 for <10%; 0 if NR."},
    {"id":"P2E_2.3","sec":"C","pri":"P2","cat":"essential","max":1,
     "param":"Processes for safe reclaim of products (e-waste, plastics, hazardous)",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P2E_2.4","sec":"C","pri":"P2","cat":"essential","max":3,
     "param":"Extended Producer Responsibility (EPR) and waste collection plan",
     "criteria":"Score 3 for EPR + waste plan in compliance + mitigation; 2 for EPR only; 1 for Reporting; 0 if NR."},
    {"id":"P2L_L2.1","sec":"C","pri":"P2","cat":"leadership","max":3,
     "param":"Life Cycle Assessments (LCA) for products/services",
     "criteria":"Score 3 if external agency + public domain; 2 if either; 1 if any; 0 if NR."},
    {"id":"P2L_L2.2","sec":"C","pri":"P2","cat":"leadership","max":3,
     "param":"Actions to mitigate environmental/social impacts per LCA",
     "criteria":"Score 3 for 100% products; 2 for 75% turnover; 1 for >50% turnover; 0 if NR."},
    {"id":"P2L_L2.3","sec":"C","pri":"P2","cat":"leadership","max":4,
     "param":"% of recycled or reused input material",
     "criteria":"Score 4 for >60%; 3 for >50%; 2 for 25-50%; 1 for <25%; 0 if NR."},
    {"id":"P2L_L2.4","sec":"C","pri":"P2","cat":"leadership","max":1,
     "param":"Quantities collected for reuse/recycling/disposal after sale",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P2L_L2.5","sec":"C","pri":"P2","cat":"leadership","max":1,
     "param":"Reclaimed products & packaging as % of total products sold",
     "criteria":"Score 1 if Reported; 0 if NR."},
    # ── PRINCIPLE 3 ──────────────────────────────────────────────────────────
    {"id":"P3E_3.1a_hi","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% employees covered under Health insurance",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1a_acc","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% employees covered by Accident insurance",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1a_mat","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% employees covered by Maternity/Paternity Benefits",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1a_day","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% employees covered by Day Care Benefits",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1b_hi","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% workmen covered under Health insurance",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1b_acc","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% workmen covered by Accident insurance",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1b_mat","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% workmen covered by Maternity/Paternity Benefits",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.1b_day","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% workmen covered by Day Care Benefits",
     "criteria":"Score 2 for >75%; 1 for >50% but ≤75%; 0 otherwise."},
    {"id":"P3E_3.3","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Premises accessible to differently abled (Rights of Persons with Disabilities Act 2016)",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.4","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Equal opportunity policy under Disabilities Act 2016",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.5","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Return to work and Retention rates after parental leave",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.6","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"Grievance redressal mechanism for employees/workmen",
     "criteria":"Score 1 for employees; 1 for workmen; 0 if neither."},
    {"id":"P3E_3.8a_hs","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% employees/workmen trained on health and safety",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P3E_3.8a_sk","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% employees/workmen trained on skill upgradation",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P3E_3.9a_emp","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% performance and career development reviews of employees",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P3E_3.9a_wkm","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% performance and career development reviews of workmen",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P3E_3.10a","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Occupational health and safety management system implemented",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.10b","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Processes to identify work-related hazards and assess risks",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.10c","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Process for workers to report hazards and remove from risk",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.10d","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Access to non-occupational medical/healthcare services",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.11","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Safety incidents causing fatalities/high-consequence injuries",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.12","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Measures to ensure safe and healthy workplace",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3E_3.13","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% complaints on working condition/health/safety resolved",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P3E_3.14","sec":"C","pri":"P3","cat":"essential","max":2,
     "param":"% plants/offices assessed for health/safety/working conditions",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P3E_3.15","sec":"C","pri":"P3","cat":"essential","max":1,
     "param":"Corrective actions on safety-related incidents and risks",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3L_L3.1","sec":"C","pri":"P3","cat":"leadership","max":1,
     "param":"Life insurance/compensatory package on death of Employee/Worker",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"P3L_L3.2","sec":"C","pri":"P3","cat":"leadership","max":2,
     "param":"Statutory dues deducted and deposited by value chain partners",
     "criteria":"Score 2 for Complied and Reported; 1 for either; 0 otherwise."},
    {"id":"P3L_L3.3","sec":"C","pri":"P3","cat":"leadership","max":3,
     "param":"Employees/workers with high-consequence injury/fatality rehabilitated",
     "criteria":"Score 3 for >80% rehabilitated; 2 for 60-80%; 1 for <60%; 0 if NR."},
    {"id":"P3L_L3.4","sec":"C","pri":"P3","cat":"leadership","max":1,
     "param":"Transition assistance programs for retirement/termination",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3L_L3.5","sec":"C","pri":"P3","cat":"leadership","max":1,
     "param":"Assessment of value chain partners – working conditions, health and safety",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P3L_L3.6","sec":"C","pri":"P3","cat":"leadership","max":1,
     "param":"Corrective actions on value chain partner H&S assessments",
     "criteria":"Score 1 if Reported and corrective actions taken; 0 if NR."},
    # ── PRINCIPLE 4 ──────────────────────────────────────────────────────────
    {"id":"P4E_4.1","sec":"C","pri":"P4","cat":"essential","max":1,
     "param":"Process for identifying key stakeholder groups",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P4E_4.2","sec":"C","pri":"P4","cat":"essential","max":4,
     "param":"Key stakeholder groups identified and frequency of engagement",
     "criteria":"Score 2 if reported for all stakeholders; 1 if not all; 0 if NR. PLUS 2 if >80% identified are vulnerable/marginalised; 1 if <80%; 0 otherwise. Max 4."},
    {"id":"P4L_L4.1","sec":"C","pri":"P4","cat":"leadership","max":1,
     "param":"Board consultation process with stakeholders on ESG",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P4L_L4.2","sec":"C","pri":"P4","cat":"leadership","max":1,
     "param":"Identification and management of ESG topics via stakeholder consultation",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P4L_L4.3","sec":"C","pri":"P4","cat":"leadership","max":3,
     "param":"Engagement with vulnerable/marginalised stakeholder groups",
     "criteria":"Score 3 for quarterly; 2 for half-yearly; 1 for annual; 0 for no engagement."},
    # ── PRINCIPLE 5 ──────────────────────────────────────────────────────────
    {"id":"P5E_5.1","sec":"C","pri":"P5","cat":"essential","max":2,
     "param":"% employees/workmen trained on human rights issues",
     "criteria":"Score 2 for 80-100%; 1 for 60-80%; 0 otherwise."},
    {"id":"P5E_5.2","sec":"C","pri":"P5","cat":"essential","max":3,
     "param":"Minimum wages paid to employees and workers",
     "criteria":"Score 3 if wages ≥30% above minimum; 2 if ≥20% above; 1 if Reported; 0 if NR."},
    {"id":"P5E_5.3","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Remuneration/salary/wages of Directors, KMP, employees, workmen",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P5E_5.4","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Focal point for addressing human rights impacts",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"P5E_5.5","sec":"C","pri":"P5","cat":"essential","max":2,
     "param":"Internal mechanisms to redress human rights grievances",
     "criteria":"Score 2 for mechanism present and reported; 1 for Reported; 0 if NR."},
    {"id":"P5E_5.6","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Sexual harassment, child labour, forced labour grievances",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P5E_5.7","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Mechanism to prevent retaliation to complainants",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P5E_5.8","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Human rights requirements in business agreements/contracts",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"P5E_5.9","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Assessment of plants/offices for child labour, forced labour, harassment",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P5E_5.10","sec":"C","pri":"P5","cat":"essential","max":1,
     "param":"Corrective actions on human rights assessment risks",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P5L_L5.1","sec":"C","pri":"P5","cat":"leadership","max":1,
     "param":"Business processes modified due to human rights grievances",
     "criteria":"Score 1 if Modified; 0 if Not modified."},
    {"id":"P5L_L5.2","sec":"C","pri":"P5","cat":"leadership","max":1,
     "param":"Scope and coverage of Human Rights due diligence (incl. value chain)",
     "criteria":"Score 1 if 80-100% covered; 0 for <60% or NR."},
    {"id":"P5L_L5.3","sec":"C","pri":"P5","cat":"leadership","max":1,
     "param":"Accessibility for differently abled visitors (Disabilities Act 2016)",
     "criteria":"Score 1 if accessibility reported; 0 if NR."},
    {"id":"P5L_L5.4","sec":"C","pri":"P5","cat":"leadership","max":2,
     "param":"Assessment of value chain partners – child/forced labour, harassment",
     "criteria":"Score 2 if 100% assessed; 1 if >75% assessed and Reported; 0 if NR."},
    {"id":"P5L_L5.5","sec":"C","pri":"P5","cat":"leadership","max":1,
     "param":"Corrective actions on value chain HR assessment risks",
     "criteria":"Score 1 if Reported; 0 if NR."},
    # ── PRINCIPLE 6 ──────────────────────────────────────────────────────────
    {"id":"P6E_6.1e","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"Total energy consumption and energy intensity vs turnover",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6E_6.1ea","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"External agency assessment of energy data",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6E_6.2","sec":"C","pri":"P6","cat":"essential","max":2,
     "param":"PAT Scheme designated consumers – targets achieved",
     "criteria":"Score 2 if targets achieved and reported; 1 if Reported; 0 if NR."},
    {"id":"P6E_6.3w","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"Water withdrawal, consumption and intensity details",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6E_6.3wa","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"External agency assessment of water data",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6E_6.4","sec":"C","pri":"P6","cat":"essential","max":2,
     "param":"Zero Liquid Discharge mechanism and coverage",
     "criteria":"Score 2 if mechanism present and reported; 1 if Reported; 0 if NR."},
    {"id":"P6E_6.5ae","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"Air emissions (NOx, SOx, PM, VOC, HAP etc.) reported",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6E_6.5aa","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"External agency assessment of air emissions",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6E_6.6g","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"GHG emissions Scope 1 and Scope 2 and intensity",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6E_6.6ga","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"External agency assessment of GHG emissions",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6E_6.7","sec":"C","pri":"P6","cat":"essential","max":5,
     "param":"Projects to reduce GHG emissions",
     "criteria":"Score 5 if projects for reducing GHG and reported; 1 if Reported; 0 if NR."},
    {"id":"P6E_6.8w","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"Waste management details (plastic, e-waste, bio-medical, etc.)",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6E_6.8wa","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"External agency assessment of waste management",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6E_6.9","sec":"C","pri":"P6","cat":"essential","max":5,
     "param":"Waste management practices and strategy to reduce hazardous chemicals",
     "criteria":"Score 5 if practices exist and reported; 1 if Reported; 0 if NR."},
    {"id":"P6E_6.10","sec":"C","pri":"P6","cat":"essential","max":1,
     "param":"Operations in ecologically sensitive areas – complying with clearances",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6E_6.11","sec":"C","pri":"P6","cat":"essential","max":4,
     "param":"Environmental impact assessments of projects",
     "criteria":"Score 4 for external assessment + public domain; 3 internal + communicated; 2 assessed not communicated; 1 Reported; 0 NR."},
    {"id":"P6E_6.12","sec":"C","pri":"P6","cat":"essential","max":5,
     "param":"Compliance with environmental laws/regulations in India",
     "criteria":"Score 5 for 100% compliance and reported; 1 if Reported; 0 if NR."},
    {"id":"P6L_L6.1","sec":"C","pri":"P6","cat":"leadership","max":2,
     "param":"Energy from renewable sources (breakdown and %)",
     "criteria":"Score 2 if reported and renewable >50%; 1 for 25-50%; 0 for <25% or NR."},
    {"id":"P6L_L6.1a","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"External agency assessment of renewable energy data",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6L_L6.2","sec":"C","pri":"P6","cat":"leadership","max":5,
     "param":"Water discharged by destination and level of treatment",
     "criteria":"Score 1 for each treatment type reported (surface/ground/seawater/3rd parties/others). Max 5; 0 if NR/no treatment."},
    {"id":"P6L_L6.2a","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"External agency assessment of water discharge",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6L_L6.3","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"Water withdrawal/consumption/discharge in water stress areas",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6L_L6.3a","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"External agency assessment of water stress data",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6L_L6.4","sec":"C","pri":"P6","cat":"leadership","max":2,
     "param":"Scope 3 GHG emissions and intensity with upstream/downstream breakup",
     "criteria":"Score 2 if reported with breakup; 0 if NR."},
    {"id":"P6L_L6.4a","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"External agency assessment of Scope 3 emissions",
     "criteria":"Score 1 if External Agency Assessment; 0 otherwise."},
    {"id":"P6L_L6.5","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"Biodiversity impact in ecologically sensitive areas + remediation",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6L_L6.6","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"Innovative technology/solutions for resource efficiency",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6L_L6.7","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"Business continuity and disaster management plan",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6L_L6.8","sec":"C","pri":"P6","cat":"leadership","max":1,
     "param":"Adverse environmental impact from value chain – mitigation methods",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P6L_L6.9","sec":"C","pri":"P6","cat":"leadership","max":2,
     "param":"% value chain partners assessed for environmental impacts",
     "criteria":"Score 2 if 100% covered and Reported; 1 if >60%; 0 for <60% or NR."},
    # ── PRINCIPLE 7 ──────────────────────────────────────────────────────────
    {"id":"P7E_7.1a","sec":"C","pri":"P7","cat":"essential","max":1,
     "param":"Affiliations with trade/industry chambers/associations (count)",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P7E_7.1b","sec":"C","pri":"P7","cat":"essential","max":1,
     "param":"Top 10 trade/industry chambers/associations listed",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P7E_7.2","sec":"C","pri":"P7","cat":"essential","max":3,
     "param":"Corrective action on adverse judicial/regulatory orders (anti-competitive)",
     "criteria":"Score 3 for no adverse order; 2 for adverse order + corrective action; 1 if Reported; 0 otherwise."},
    {"id":"P7L_L7.1","sec":"C","pri":"P7","cat":"leadership","max":2,
     "param":"Public policy positions advocated by the company",
     "criteria":"Score 2 for Positions advocated; 0 otherwise."},
    # ── PRINCIPLE 8 ──────────────────────────────────────────────────────────
    {"id":"P8E_8.1","sec":"C","pri":"P8","cat":"essential","max":5,
     "param":"Social Impact Assessments (SIA) conducted",
     "criteria":"Score 5 for external agency; 2 for internal; 0 if NR."},
    {"id":"P8E_8.2","sec":"C","pri":"P8","cat":"essential","max":3,
     "param":"Ongoing Rehabilitation and Resettlement projects",
     "criteria":"Score 3 for >2 projects; 2 for 2 projects; 1 for one; 0 if NR."},
    {"id":"P8E_8.3","sec":"C","pri":"P8","cat":"essential","max":1,
     "param":"Mechanisms to receive/redress community grievances",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P8E_8.4","sec":"C","pri":"P8","cat":"essential","max":6,
     "param":"% inputs sourced from MSMEs / local suppliers within district",
     "criteria":"Score 3 for MSME/small producers >80%; 2 for >50%; PLUS 3 for >50% local; 2 for <50%; 1 for Reported; 0 NR. Max 6."},
    {"id":"P8L_L8.1","sec":"C","pri":"P8","cat":"leadership","max":1,
     "param":"Actions to mitigate negative social impacts from SIA",
     "criteria":"Score 1 for corrective actions for all activities; 0 otherwise."},
    {"id":"P8L_L8.2","sec":"C","pri":"P8","cat":"leadership","max":2,
     "param":"CSR in government-designated aspirational districts",
     "criteria":"Score 2 for all designated districts; 0 if NR."},
    {"id":"P8L_L8.3","sec":"C","pri":"P8","cat":"leadership","max":1,
     "param":"Preference for purchasing from marginal/vulnerable groups",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"P8L_L8.3p","sec":"C","pri":"P8","cat":"leadership","max":2,
     "param":"% of total procurement from marginal/vulnerable suppliers",
     "criteria":"Score 2 for >20%; 1 for 10-20%; 0 for <10% or NR."},
    {"id":"P8L_L8.4","sec":"C","pri":"P8","cat":"leadership","max":1,
     "param":"Basis for sharing benefits from intellectual properties",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P8L_L8.5","sec":"C","pri":"P8","cat":"leadership","max":1,
     "param":"Corrective actions in IP-related cases",
     "criteria":"Score 1 for corrective actions for all IP cases; 0 otherwise."},
    {"id":"P8L_L8.6","sec":"C","pri":"P8","cat":"leadership","max":2,
     "param":"CSR beneficiaries – % of vulnerable/marginalised groups",
     "criteria":"Score 2 for >80%; 1 for 50-80%; 0 for <50% or NR."},
    # ── PRINCIPLE 9 ──────────────────────────────────────────────────────────
    {"id":"P9E_9.1","sec":"C","pri":"P9","cat":"essential","max":2,
     "param":"Mechanism to receive/respond to consumer complaints and feedback",
     "criteria":"Score 2 for mechanism present and reported; 1 if Reported; 0 if NR."},
    {"id":"P9E_9.2e","sec":"C","pri":"P9","cat":"essential","max":5,
     "param":"% products/services with environmental and social parameter info",
     "criteria":"Score 5 for 90-100%; 4 for 75-90%; 3 for 70-75%; 2 for 60-70%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.2s","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% products/services with safe/responsible usage information",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.2r","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% products/services with recycling and safe disposal information",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.3dp","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% consumer complaints on data privacy resolved",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR. Max score if no complaints."},
    {"id":"P9E_9.3ad","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% consumer complaints on advertising resolved",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.3cs","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% consumer complaints on cyber security resolved",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.3es","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% consumer complaints on essential services resolved",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.3rt","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% consumer complaints on restrictive trade practices resolved",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.3ut","sec":"C","pri":"P9","cat":"essential","max":3,
     "param":"% consumer complaints on unfair trade practices resolved",
     "criteria":"Score 3 for 80-100%; 2 for 60-80%; 1 for <60%; 0 NR."},
    {"id":"P9E_9.4","sec":"C","pri":"P9","cat":"essential","max":2,
     "param":"Product recall instances due to safety issues",
     "criteria":"Score 2 if Reported and no recalls; 1 if Reported; 0 if NR. Max if no recalls."},
    {"id":"P9E_9.5","sec":"C","pri":"P9","cat":"essential","max":2,
     "param":"Cyber security framework/policy and data privacy risks",
     "criteria":"Score 2 if framework exists and reported; 1 if Reported; 0 if NR."},
    {"id":"P9E_9.6","sec":"C","pri":"P9","cat":"essential","max":1,
     "param":"Corrective actions on data privacy/advertising complaints",
     "criteria":"Score 1 if steps taken; 0 if NR."},
    {"id":"P9L_L9.1","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"Channels/platforms for product/service information access",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P9L_L9.2","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"Steps to educate consumers on safe/responsible usage",
     "criteria":"Score 1 if steps taken; 0 if NR."},
    {"id":"P9L_L9.3","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"Mechanism to inform consumers of service disruption risk",
     "criteria":"Score 1 for Yes; 0 for No/NR."},
    {"id":"P9L_L9.4p","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"Product information beyond legal mandate",
     "criteria":"Score 1 for Yes/NA; 0 for No/NR."},
    {"id":"P9L_L9.4s","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"Consumer survey conducted by the company",
     "criteria":"Score 1 for Yes; 0 for No."},
    {"id":"P9L_L9.5a","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"Data breach instances and impact reported",
     "criteria":"Score 1 if Reported; 0 if NR."},
    {"id":"P9L_L9.5b","sec":"C","pri":"P9","cat":"leadership","max":1,
     "param":"% data breaches involving PII of customers",
     "criteria":"Score 1 if Reported; 0 if NR."},
]

# Build lookup: max scores per section/principle
SECTION_MAX = {"A": 18, "B": 24}
PRINCIPLE_MAX = {"P1":24,"P2":27,"P3":49,"P4":10,"P5":20,"P6":54,"P7":7,"P8":24,"P9":43}
GRAND_MAX = 300

# ═══════════════════════════════════════════════════════════════════════════════
# VERTEX AI SETUP
# ═══════════════════════════════════════════════════════════════════════════════
_model = None
_lock  = threading.Lock()

def get_model() -> GenerativeModel:
    global _model
    if _model is None:
        with _lock:
            if _model is None:
                creds = service_account.Credentials.from_service_account_file(
                    str(KEY_FILE),
                    scopes=["https://www.googleapis.com/auth/cloud-platform"],
                )
                vertexai.init(project=PROJECT_ID, location=LOCATION, credentials=creds)
                _model = GenerativeModel(MODEL_NAME)
                log.info(f"Vertex AI initialized  project={PROJECT_ID}  model={MODEL_NAME}")
    return _model

# ═══════════════════════════════════════════════════════════════════════════════
# PROMPT BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def build_prompt(company_name: str) -> str:
    indicators_text = "\n".join(
        f'  "{q["id"]}": {{"param": "{q["param"]}", "max_score": {q["max"]}, '
        f'"criteria": "{q["criteria"]}"}}'
        for q in SRMM_SCHEMA
    )
    return f"""You are an expert BRSR (Business Responsibility and Sustainability Report) analyst
specializing in the SRMM (Sustainability Reporting Maturity Model) v2.0 framework by ICAI.

Analyze the attached BRSR report for: **{company_name}**

For each indicator below, read the relevant section of the report and:
1. Determine the score based on the exact criteria given
2. Write a brief reason (1-2 sentences referencing the report)
3. Cite the evidence (page/section reference if possible)

If information is NOT reported (NR), score = 0 and say "Not Reported".

Indicators to score:
{{
{indicators_text}
}}

Return ONLY valid JSON in EXACTLY this structure (no markdown, no extra text):
{{
  "indicators": {{
    "<indicator_id>": {{
      "score": <integer>,
      "max_score": <integer>,
      "reason": "<string>",
      "evidence": "<string>"
    }},
    ...
  }}
}}
"""

# ═══════════════════════════════════════════════════════════════════════════════
# PDF DOWNLOAD
# ═══════════════════════════════════════════════════════════════════════════════
def download_pdf(url: str) -> bytes | None:
    try:
        r = requests.get(url, timeout=PDF_DOWNLOAD_TIMEOUT, stream=True)
        r.raise_for_status()
        data = b"".join(r.iter_content(chunk_size=65536))
        if len(data) < 1000:
            log.warning(f"PDF too small ({len(data)} bytes): {url}")
            return None
        return data
    except Exception as e:
        log.warning(f"PDF download failed: {url}  →  {e}")
        return None

# ═══════════════════════════════════════════════════════════════════════════════
# GEMINI ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
def analyze_with_gemini(pdf_bytes: bytes, company_name: str) -> dict | None:
    model  = get_model()
    prompt = build_prompt(company_name)
    pdf_part = Part.from_data(data=pdf_bytes, mime_type="application/pdf")
    cfg = GenerationConfig(
        temperature=0.1,
        response_mime_type="application/json",
    )
    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            resp = model.generate_content(
                [pdf_part, prompt],
                generation_config=cfg,
                stream=False,
            )
            raw = resp.text.strip()
            # strip markdown fences if any
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw)
            parsed = json.loads(raw)
            if "indicators" not in parsed:
                raise ValueError("Missing 'indicators' key in response")
            return parsed["indicators"]
        except json.JSONDecodeError as e:
            log.warning(f"[{company_name}] JSON parse error attempt {attempt}: {e}")
        except Exception as e:
            log.warning(f"[{company_name}] Gemini error attempt {attempt}: {e}")
        if attempt < RETRY_ATTEMPTS:
            delay = RETRY_DELAY_BASE * (2 ** (attempt - 1))
            log.info(f"[{company_name}] Retrying in {delay}s …")
            time.sleep(delay)
    return None

# ═══════════════════════════════════════════════════════════════════════════════
# SCORE CALCULATION
# ═══════════════════════════════════════════════════════════════════════════════
def calculate_scores(indicators: dict) -> dict:
    section_a = section_b = 0
    principles = {f"P{i}": {"essential": 0, "leadership": 0, "total": 0,
                             "max_essential": 0, "max_leadership": 0, "max": 0}
                  for i in range(1, 10)}

    for q in SRMM_SCHEMA:
        qid  = q["id"]
        sec  = q["sec"]
        cat  = q["cat"]
        pri  = q.get("pri", "")
        mmax = q["max"]
        raw  = indicators.get(qid, {})
        sc   = min(int(raw.get("score", 0)), mmax)  # clamp to max

        if sec == "A":
            section_a += sc
        elif sec == "B":
            section_b += sc
        elif sec == "C" and pri in principles:
            if cat == "essential":
                principles[pri]["essential"] += sc
                principles[pri]["max_essential"] += mmax
            else:
                principles[pri]["leadership"] += sc
                principles[pri]["max_leadership"] += mmax
            principles[pri]["total"] += sc
            principles[pri]["max"]   += mmax

    # pre-fill max from schema (so missing companies still have correct max)
    for p, v in principles.items():
        v["max"] = PRINCIPLE_MAX[p]

    essential_total   = section_a + section_b + sum(v["essential"] for v in principles.values())
    leadership_total  = sum(v["leadership"] for v in principles.values())
    grand_total       = essential_total + leadership_total
    pct               = round(grand_total / GRAND_MAX * 100, 2)

    if pct <= 25:
        level, stage = "Level 1", "Formative Stage"
    elif pct <= 50:
        level, stage = "Level 2", "Emerging Stage"
    elif pct <= 75:
        level, stage = "Level 3", "Established Stage"
    else:
        level, stage = "Level 4", "Leading by Example"

    return {
        "section_a":        section_a,
        "section_a_max":    18,
        "section_b":        section_b,
        "section_b_max":    24,
        "principles":       principles,
        "essential_total":  essential_total,
        "leadership_total": leadership_total,
        "grand_total":      grand_total,
        "max_score":        GRAND_MAX,
        "percentage":       pct,
        "maturity_level":   level,
        "maturity_stage":   stage,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# PROGRESS PERSISTENCE
# ═══════════════════════════════════════════════════════════════════════════════
_results_lock = threading.Lock()
_shutdown_requested = threading.Event()

def _handle_signal(signum, frame):
    log.warning(f"Signal {signum} received — finishing in-flight requests then saving & exiting …")
    _shutdown_requested.set()

signal.signal(signal.SIGTERM, _handle_signal)
signal.signal(signal.SIGINT,  _handle_signal)

def load_progress() -> dict:
    """Returns {sno: result_dict} of already-completed companies."""
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_progress(done: dict):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(done, f, ensure_ascii=False)

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESS ONE COMPANY
# ═══════════════════════════════════════════════════════════════════════════════
def process_company(company: dict) -> dict:
    sno  = company["sno"]
    name = company["company"]
    url  = company["brsr_link"]
    ts   = datetime.utcnow().isoformat()

    base = {"sno": sno, "company": name, "brsr_link": url, "analyzed_at": ts}

    if not url:
        log.info(f"[{sno:4d}] {name}  →  no link, skipping")
        return {**base, "status": "no_link", "answers": {}, "scores": None}

    log.info(f"[{sno:4d}] {name}  →  downloading PDF …")
    pdf = download_pdf(url)
    if pdf is None:
        return {**base, "status": "pdf_download_failed", "answers": {}, "scores": None}

    log.info(f"[{sno:4d}] {name}  →  analysing with Gemini ({len(pdf)//1024} KB) …")
    indicators = analyze_with_gemini(pdf, name)
    if indicators is None:
        return {**base, "status": "gemini_failed", "answers": {}, "scores": None}

    scores = calculate_scores(indicators)
    log.info(
        f"[{sno:4d}] {name}  →  DONE  "
        f"{scores['grand_total']}/{GRAND_MAX} ({scores['percentage']}%)  "
        f"{scores['maturity_level']}"
    )
    return {**base, "status": "completed", "answers": indicators, "scores": scores}

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
LEVEL_COLORS = {
    "Level 1": "FF4C4C",   # red
    "Level 2": "FFA500",   # orange
    "Level 3": "FFD700",   # yellow/gold
    "Level 4": "4CAF50",   # green
    "N/A":     "CCCCCC",   # grey
}
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=9)
WRAP      = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="center")

def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill  = HDR_FILL
    c.font  = HDR_FONT
    c.alignment = CENTER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def save_excel(all_results: list):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    hdrs = ["S.No","Company","Status","Section A","Section B",
            "P1","P2","P3","P4","P5","P6","P7","P8","P9",
            "Essential","Leadership","Grand Total","Max","% Score","Maturity Level","Stage"]
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws1, 1, ci, h, width=max(8, len(h)+2))
    ws1.column_dimensions["B"].width = 45
    ws1.column_dimensions["U"].width = 18
    ws1.freeze_panes = "A2"

    for ri, r in enumerate(all_results, 2):
        sc = r.get("scores") or {}
        pr = sc.get("principles", {})
        level = sc.get("maturity_level", "N/A") if sc else "N/A"
        fill  = PatternFill("solid", fgColor=LEVEL_COLORS.get(level, "CCCCCC"))
        row_data = [
            r["sno"], r["company"], r["status"],
            sc.get("section_a",""),    sc.get("section_b",""),
            pr.get("P1",{}).get("total",""), pr.get("P2",{}).get("total",""),
            pr.get("P3",{}).get("total",""), pr.get("P4",{}).get("total",""),
            pr.get("P5",{}).get("total",""), pr.get("P6",{}).get("total",""),
            pr.get("P7",{}).get("total",""), pr.get("P8",{}).get("total",""),
            pr.get("P9",{}).get("total",""),
            sc.get("essential_total",""),  sc.get("leadership_total",""),
            sc.get("grand_total",""),      sc.get("max_score",""),
            sc.get("percentage",""),       level,
            sc.get("maturity_stage",""),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            if ci >= 20:
                c.fill = fill
                c.alignment = CENTER

    # ── Sheet 2: Detailed indicator scores ───────────────────────────────────
    ws2 = wb.create_sheet("Indicator Scores")
    fixed = ["S.No","Company","Status"]
    ind_ids = [q["id"] for q in SRMM_SCHEMA]
    ind_params = {q["id"]: q["param"] for q in SRMM_SCHEMA}
    ind_max    = {q["id"]: q["max"]   for q in SRMM_SCHEMA}
    for ci, h in enumerate(fixed + ind_ids, 1):
        label = h if h in fixed else f"{h}\n(max {ind_max.get(h,'')})"
        _hdr(ws2, 1, ci, label, width=14 if h not in fixed else (6 if h=="S.No" else 45))
    ws2.row_dimensions[1].height = 40
    ws2.freeze_panes = "D2"

    for ri, r in enumerate(all_results, 2):
        ans = r.get("answers") or {}
        row_data = [r["sno"], r["company"], r["status"]]
        for qid in ind_ids:
            row_data.append(ans.get(qid, {}).get("score", ""))
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT

    # ── Sheet 3: Answers (reason + evidence) ─────────────────────────────────
    ws3 = wb.create_sheet("Answers")
    ans_hdrs = ["S.No","Company","Indicator ID","Parameter","Score","Max","Reason","Evidence"]
    for ci, h in enumerate(ans_hdrs, 1):
        _hdr(ws3, 1, ci, h, width=14)
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["D"].width = 45
    ws3.column_dimensions["G"].width = 55
    ws3.column_dimensions["H"].width = 45
    ws3.freeze_panes = "A2"

    row_idx = 2
    for r in all_results:
        if r["status"] != "completed":
            continue
        ans = r.get("answers") or {}
        for qid in ind_ids:
            a = ans.get(qid, {})
            ws3.cell(row=row_idx, column=1, value=r["sno"]).font = BODY_FONT
            ws3.cell(row=row_idx, column=2, value=r["company"]).font = BODY_FONT
            ws3.cell(row=row_idx, column=3, value=qid).font = BODY_FONT
            ws3.cell(row=row_idx, column=4, value=ind_params.get(qid,"")).font = BODY_FONT
            ws3.cell(row=row_idx, column=5, value=a.get("score","")).font = BODY_FONT
            ws3.cell(row=row_idx, column=6, value=ind_max.get(qid,"")).font = BODY_FONT
            c7 = ws3.cell(row=row_idx, column=7, value=a.get("reason",""))
            c7.font = BODY_FONT; c7.alignment = WRAP
            c8 = ws3.cell(row=row_idx, column=8, value=a.get("evidence",""))
            c8.font = BODY_FONT; c8.alignment = WRAP
            row_idx += 1

    wb.save(OUTPUT_EXCEL)
    log.info(f"Excel saved → {OUTPUT_EXCEL}")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    log.info("══════════════════════════════════════════════════")
    log.info("  SRMM Scraper  —  Gemini 2.5 Pro  —  starting   ")
    log.info("══════════════════════════════════════════════════")

    # Load companies
    with open(COMPANIES_JSON) as f:
        companies = json.load(f)
    log.info(f"Loaded {len(companies)} companies from {COMPANIES_JSON}")

    # Load progress
    done = load_progress()
    log.info(f"Already completed: {len(done)} companies")

    pending = [c for c in companies if str(c["sno"]) not in done]
    log.info(f"Pending: {len(pending)} companies")

    # Warm up Vertex AI
    get_model()

    # Process companies (parallel)
    def _flush(done, companies):
        """Save progress + incremental JSON snapshot."""
        save_progress(done)
        all_res = [done[str(co["sno"])] for co in companies if str(co["sno"]) in done]
        out = {
            "metadata": {
                "generated_at": datetime.utcnow().isoformat(),
                "total_companies": len(companies),
                "completed": sum(1 for v in done.values() if v.get("status") == "completed"),
                "no_link":   sum(1 for v in done.values() if v.get("status") == "no_link"),
                "failed":    sum(1 for v in done.values() if v.get("status") in
                                 ("gemini_failed","pdf_download_failed","error")),
                "model": MODEL_NAME,
            },
            "companies": all_res,
        }
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(out, f, indent=2, ensure_ascii=False)
        return all_res

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(process_company, c): c for c in pending}
        try:
            for fut in as_completed(futures):
                c = futures[fut]
                try:
                    result = fut.result()
                except Exception as e:
                    log.error(f"Unexpected error for sno={c['sno']}: {e}\n{traceback.format_exc()}")
                    result = {
                        "sno": c["sno"], "company": c["company"],
                        "brsr_link": c["brsr_link"], "status": "error",
                        "error": str(e), "answers": {}, "scores": None,
                        "analyzed_at": datetime.utcnow().isoformat(),
                    }
                with _results_lock:
                    done[str(c["sno"])] = result
                    _flush(done, companies)

                if _shutdown_requested.is_set():
                    log.warning("Shutdown requested — cancelling remaining futures …")
                    for f in futures:
                        f.cancel()
                    break
        except Exception as e:
            log.error(f"Unexpected loop error: {e}\n{traceback.format_exc()}")
        finally:
            # Always save whatever we have before exiting
            with _results_lock:
                _flush(done, companies)
            log.info(f"Progress saved — {len(done)} companies done so far.")

    # Final saves
    if _shutdown_requested.is_set():
        log.info("Scraper stopped early. Re-run to resume from where it left off.")
    else:
        log.info("All companies processed. Writing final outputs …")
    all_results = [done.get(str(c["sno"]), {
        "sno": c["sno"], "company": c["company"],
        "brsr_link": c["brsr_link"], "status": "not_processed",
        "answers": {}, "scores": None,
    }) for c in companies]

    final_out = {
        "metadata": {
            "generated_at": datetime.utcnow().isoformat(),
            "total_companies": len(companies),
            "completed": sum(1 for r in all_results if r.get("status") == "completed"),
            "no_link": sum(1 for r in all_results if r.get("status") == "no_link"),
            "failed": sum(1 for r in all_results if r.get("status") in
                         ("gemini_failed","pdf_download_failed","error")),
            "model": MODEL_NAME,
            "srmm_version": "2.0",
            "max_possible_score": GRAND_MAX,
        },
        "companies": all_results,
    }
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(final_out, f, indent=2, ensure_ascii=False)
    log.info(f"JSON saved → {OUTPUT_JSON}")

    save_excel(all_results)

    log.info("══════════════════════════════════════════════════")
    log.info(f"  Done!  Completed: {final_out['metadata']['completed']}  "
             f"No link: {final_out['metadata']['no_link']}  "
             f"Failed: {final_out['metadata']['failed']}")
    log.info("══════════════════════════════════════════════════")

if __name__ == "__main__":
    main()
